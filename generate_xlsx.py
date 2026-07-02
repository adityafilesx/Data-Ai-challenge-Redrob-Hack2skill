#!/usr/bin/env python3
"""
generate_xlsx.py — Ranked Candidate Output (XLSX)

Reads the ranked CSV (team_lyle.csv), ranking explanations (ranking_explanation.json),
and candidate source profiles (candidates.jsonl / sample_candidates.json) to produce
a richly formatted XLSX workbook with two sheets:

  Sheet 1 — "Ranked Candidates"    : Recruiter-ready overview with scores & reasoning
  Sheet 2 — "Score Breakdown"      : Per-component scoring detail for transparency

Usage:
    pip install openpyxl
    python generate_xlsx.py [--candidates candidates.jsonl] [--output ranked_candidates.xlsx]
"""

import argparse
import csv
import gzip
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
except ImportError:
    print("ERROR: openpyxl is required. Install it with:  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Colour Palette ──────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")   # Deep navy
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2D4F7C", end_color="2D4F7C", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="E8EDF3", size=10)

ROW_EVEN_FILL = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")   # Soft blue-grey
ROW_ODD_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

TITLE_FONT    = Font(name="Calibri", bold=True, size=14, color="1B2A4A")
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="6B7A8D")
BODY_FONT     = Font(name="Calibri", size=10, color="2C3E50")
SCORE_FONT    = Font(name="Calibri", bold=True, size=11, color="1B6B3A")
RANK_FONT     = Font(name="Calibri", bold=True, size=12, color="1B2A4A")
LINK_FONT     = Font(name="Calibri", size=10, color="2E86C1", underline="single")

GOLD_FILL     = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")    # Top 10 highlight
SILVER_FILL   = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")    # Top 11-25

THIN_BORDER   = Border(
    left=Side(style="thin", color="D5DDE5"),
    right=Side(style="thin", color="D5DDE5"),
    top=Side(style="thin", color="D5DDE5"),
    bottom=Side(style="thin", color="D5DDE5"),
)

GREEN_FILL    = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
YELLOW_FILL   = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
RED_FILL      = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")


# ── Data Loaders ────────────────────────────────────────────────────────────
def load_candidates_map(path):
    """Load candidate profiles into a dict keyed by candidate_id."""
    p = Path(path)
    records = []
    if p.suffix == '.json':
        with open(p, 'r', encoding='utf-8') as f:
            data = json.load(f)
            records = data if isinstance(data, list) else [data]
    else:
        opener = gzip.open if p.suffix == '.gz' else open
        mode = 'rt' if p.suffix == '.gz' else 'r'
        with opener(p, mode, encoding='utf-8') as f:
            records = [json.loads(line) for line in f if line.strip()]
    return {r['candidate_id']: r for r in records}


def load_ranking_csv(path):
    """Load ranked output CSV."""
    rows = []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def load_explanations(path):
    """Load ranking_explanation.json."""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ── Helper: Extract summary fields from candidate ──────────────────────────
def extract_profile_fields(cand):
    """Return a flat dict of useful profile fields."""
    p = cand.get('profile', {})
    rs = cand.get('redrob_signals', {})
    edu = cand.get('education', [])
    skills = cand.get('skills', [])
    certs = cand.get('certifications', [])

    # Top skills (by endorsements, descending)
    sorted_skills = sorted(skills, key=lambda s: s.get('endorsements', 0), reverse=True)
    top_skill_names = [s['name'] for s in sorted_skills[:6]]

    # Education summary
    edu_summary = ""
    if edu:
        top_edu = edu[0]
        edu_summary = f"{top_edu.get('degree', '')} {top_edu.get('field_of_study', '')} — {top_edu.get('institution', '')}".strip()

    # Certifications
    cert_names = [c.get('name', '') for c in certs[:3]]

    # Assessment scores
    assessments = rs.get('skill_assessment_scores', {})
    assessment_str = "; ".join(f"{k}: {v}" for k, v in assessments.items()) if assessments else "None"

    salary = rs.get('expected_salary_range_inr_lpa', {})
    salary_str = f"₹{salary.get('min', '?')}–{salary.get('max', '?')} LPA" if salary else "N/A"

    return {
        'name': p.get('anonymized_name', 'Unknown'),
        'headline': p.get('headline', ''),
        'current_title': p.get('current_title', ''),
        'current_company': p.get('current_company', ''),
        'current_industry': p.get('current_industry', ''),
        'location': f"{p.get('location', '')}, {p.get('country', '')}".strip(', '),
        'years_of_experience': p.get('years_of_experience', 0),
        'top_skills': ', '.join(top_skill_names),
        'education': edu_summary,
        'certifications': ', '.join(cert_names) if cert_names else 'None',
        'assessment_scores': assessment_str,
        'salary_range': salary_str,
        'notice_period': rs.get('notice_period_days', 'N/A'),
        'open_to_work': '✅ Yes' if rs.get('open_to_work_flag') else '❌ No',
        'work_mode': rs.get('preferred_work_mode', 'N/A'),
        'willing_to_relocate': '✅ Yes' if rs.get('willing_to_relocate') else '❌ No',
        'github_score': rs.get('github_activity_score', -1),
        'profile_completeness': rs.get('profile_completeness_score', 0),
        'response_rate': rs.get('recruiter_response_rate', 0),
    }


# ── Sheet 1: Ranked Candidates ─────────────────────────────────────────────
def build_ranked_sheet(ws, ranked_rows, explanations, cand_map):
    """Build the primary recruiter-facing ranked candidates sheet."""

    # Title row
    ws.merge_cells('A1:R1')
    ws['A1'] = "🏆  Ranked Candidate Recommendations — Redrob AI Challenge"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:R2')
    ws['A2'] = f"Top {len(ranked_rows)} candidates ranked by multi-component technical fit score  ·  Generated by Team Lyle scoring pipeline"
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 20

    # Header row (row 4)
    headers = [
        ("Rank", 7),
        ("Candidate ID", 16),
        ("Name", 18),
        ("Final Score", 13),
        ("Current Title", 22),
        ("Company", 18),
        ("Industry", 16),
        ("Location", 18),
        ("YoE", 7),
        ("Top Skills", 40),
        ("Education", 32),
        ("Assessments", 30),
        ("Salary Range", 16),
        ("Notice (days)", 13),
        ("Open to Work", 13),
        ("Work Mode", 12),
        ("GitHub Score", 12),
        ("Reasoning", 65),
    ]

    header_row = 4
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 30

    # Freeze panes: freeze header and rank/name columns
    ws.freeze_panes = 'D5'

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:R{header_row + len(ranked_rows)}"

    # Data rows
    for row_offset, row_data in enumerate(ranked_rows):
        data_row = header_row + 1 + row_offset
        cid = row_data.get('candidate_id', '')
        rank = int(row_data.get('rank', row_offset + 1))
        score = float(row_data.get('score', 0))
        reasoning = row_data.get('reasoning', '')

        # Get full profile data
        cand = cand_map.get(cid, {})
        pf = extract_profile_fields(cand) if cand else {}

        github = pf.get('github_score', -1)
        github_display = github if github >= 0 else "N/A"

        values = [
            rank,
            cid,
            pf.get('name', 'Unknown'),
            round(score, 2),
            pf.get('current_title', ''),
            pf.get('current_company', ''),
            pf.get('current_industry', ''),
            pf.get('location', ''),
            pf.get('years_of_experience', ''),
            pf.get('top_skills', ''),
            pf.get('education', ''),
            pf.get('assessment_scores', ''),
            pf.get('salary_range', ''),
            pf.get('notice_period', ''),
            pf.get('open_to_work', ''),
            pf.get('work_mode', ''),
            github_display,
            reasoning,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx >= 10))

        # Style: Rank column
        ws.cell(row=data_row, column=1).font = RANK_FONT
        ws.cell(row=data_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Style: Score column
        ws.cell(row=data_row, column=4).font = SCORE_FONT
        ws.cell(row=data_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=data_row, column=4).number_format = '0.00'

        # Alternating row fill + tier-based highlighting
        if rank <= 10:
            fill = GOLD_FILL
        elif rank <= 25:
            fill = SILVER_FILL
        elif row_offset % 2 == 0:
            fill = ROW_EVEN_FILL
        else:
            fill = ROW_ODD_FILL

        for col_idx in range(1, len(values) + 1):
            ws.cell(row=data_row, column=col_idx).fill = fill

        ws.row_dimensions[data_row].height = 38

    # ── Conditional formatting: Data bars on score column ───────────────
    score_range = f"D{header_row+1}:D{header_row + len(ranked_rows)}"
    ws.conditional_formatting.add(
        score_range,
        DataBarRule(start_type='min', end_type='max', color="2E86C1")
    )


# ── Sheet 2: Score Breakdown ───────────────────────────────────────────────
def build_breakdown_sheet(ws, ranked_rows, explanations, cand_map):
    """Build the detailed scoring breakdown sheet."""

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "📊  Score Breakdown — Component-Level Detail"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:L2')
    ws['A2'] = "Transparency layer: see exactly how each candidate's final score was composed"
    ws['A2'].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 20

    # Headers
    breakdown_headers = [
        ("Rank", 7),
        ("Candidate ID", 16),
        ("Name", 18),
        ("Final Score", 13),
        ("Title Score", 12),
        ("Skills Score", 13),
        ("Assessment Boost", 14),
        ("Technical Score", 14),
        ("Experience Bonus", 14),
        ("Education Bonus", 14),
        ("Location Bonus", 13),
        ("Behavioral Mult.", 14),
    ]

    header_row = 4
    for col_idx, (header, width) in enumerate(breakdown_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = 'D5'
    ws.auto_filter.ref = f"A{header_row}:L{header_row + len(ranked_rows)}"

    for row_offset, row_data in enumerate(ranked_rows):
        data_row = header_row + 1 + row_offset
        cid = row_data.get('candidate_id', '')
        rank = int(row_data.get('rank', row_offset + 1))
        score = float(row_data.get('score', 0))

        cand = cand_map.get(cid, {})
        name = cand.get('profile', {}).get('anonymized_name', 'Unknown') if cand else 'Unknown'

        exp = explanations.get(cid, {})

        values = [
            rank,
            cid,
            name,
            round(score, 2),
            round(exp.get('title_score', 0), 2),
            round(exp.get('skills_score', 0), 2),
            round(exp.get('assessment_score', 1.0), 2),
            round(exp.get('technical_score', 0), 2),
            round(exp.get('experience_bonus', 0), 2),
            round(exp.get('education_bonus', 0), 2),
            round(exp.get('location_bonus', 0), 2),
            round(exp.get('behavior_score', 0), 2),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Rank styling
        ws.cell(row=data_row, column=1).font = RANK_FONT
        # Score styling
        ws.cell(row=data_row, column=4).font = SCORE_FONT
        ws.cell(row=data_row, column=4).number_format = '0.00'

        # Number formats for component scores
        for col_idx in range(5, 13):
            ws.cell(row=data_row, column=col_idx).number_format = '0.00'

        # Row fill
        if rank <= 10:
            fill = GOLD_FILL
        elif rank <= 25:
            fill = SILVER_FILL
        elif row_offset % 2 == 0:
            fill = ROW_EVEN_FILL
        else:
            fill = ROW_ODD_FILL

        for col_idx in range(1, len(values) + 1):
            ws.cell(row=data_row, column=col_idx).fill = fill

        ws.row_dimensions[data_row].height = 24

    # ── Conditional formatting: colour-code behavioral multiplier ──────
    bm_range = f"L{header_row+1}:L{header_row + len(ranked_rows)}"
    ws.conditional_formatting.add(bm_range, CellIsRule(operator='greaterThanOrEqual', formula=['1.0'], fill=GREEN_FILL))
    ws.conditional_formatting.add(bm_range, CellIsRule(operator='between', formula=['0.8', '0.99'], fill=YELLOW_FILL))
    ws.conditional_formatting.add(bm_range, CellIsRule(operator='lessThan', formula=['0.8'], fill=RED_FILL))

    # Data bars on Final Score
    score_range = f"D{header_row+1}:D{header_row + len(ranked_rows)}"
    ws.conditional_formatting.add(
        score_range,
        DataBarRule(start_type='min', end_type='max', color="27AE60")
    )

    # ── Bar Chart: Top 20 Final Scores ──────────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.title = "Top 20 — Final Scores"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Rank"
    chart.style = 10

    num_chart_rows = min(20, len(ranked_rows))
    data_ref = Reference(ws, min_col=4, min_row=header_row, max_row=header_row + num_chart_rows)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + num_chart_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 28
    chart.height = 14

    # Colour the bars
    series = chart.series[0]
    series.graphicalProperties.solidFill = "2E86C1"

    chart_anchor_row = header_row + len(ranked_rows) + 3
    ws.add_chart(chart, f"A{chart_anchor_row}")


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate ranked candidate XLSX report")
    parser.add_argument('--candidates', default=None,
                        help="Path to candidates.jsonl or sample_candidates.json")
    parser.add_argument('--ranking-csv', default='team_lyle.csv',
                        help="Path to ranked output CSV")
    parser.add_argument('--explanations', default='ranking_explanation.json',
                        help="Path to ranking_explanation.json")
    parser.add_argument('--output', default='ranked_candidates.xlsx',
                        help="Output XLSX filename")
    args = parser.parse_args()

    base = Path(__file__).parent

    # Resolve candidate source
    if args.candidates:
        cand_path = Path(args.candidates)
    elif (base / 'candidates.jsonl').exists():
        cand_path = base / 'candidates.jsonl'
    elif (base / 'sample_candidates.json').exists():
        cand_path = base / 'sample_candidates.json'
    else:
        print("WARNING: No candidate source found. XLSX will have limited profile data.", file=sys.stderr)
        cand_path = None

    # Load data
    csv_path = base / args.ranking_csv if not Path(args.ranking_csv).is_absolute() else Path(args.ranking_csv)
    exp_path = base / args.explanations if not Path(args.explanations).is_absolute() else Path(args.explanations)

    print("📂 Loading ranked CSV...", file=sys.stderr)
    ranked_rows = load_ranking_csv(csv_path)

    print("📂 Loading explanations...", file=sys.stderr)
    explanations = load_explanations(exp_path)

    cand_map = {}
    if cand_path and cand_path.exists():
        print(f"📂 Loading candidate profiles from {cand_path.name}...", file=sys.stderr)
        cand_map = load_candidates_map(cand_path)
        print(f"   Loaded {len(cand_map):,} profiles.", file=sys.stderr)
    else:
        print("⚠️  No candidate profiles loaded — XLSX will have limited data.", file=sys.stderr)

    # Build workbook
    wb = Workbook()

    # Sheet 1: Ranked Candidates
    ws1 = wb.active
    ws1.title = "Ranked Candidates"
    ws1.sheet_properties.tabColor = "1B2A4A"
    print("🔨 Building Sheet 1: Ranked Candidates...", file=sys.stderr)
    build_ranked_sheet(ws1, ranked_rows, explanations, cand_map)

    # Sheet 2: Score Breakdown
    ws2 = wb.create_sheet(title="Score Breakdown")
    ws2.sheet_properties.tabColor = "27AE60"
    print("🔨 Building Sheet 2: Score Breakdown...", file=sys.stderr)
    build_breakdown_sheet(ws2, ranked_rows, explanations, cand_map)

    # Save
    out_path = base / args.output if not Path(args.output).is_absolute() else Path(args.output)
    wb.save(out_path)
    print(f"\n✅ XLSX written to: {out_path}", file=sys.stderr)
    print(f"   Rows: {len(ranked_rows)}  |  Sheets: {wb.sheetnames}", file=sys.stderr)


if __name__ == '__main__':
    main()
